import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
import xgboost as xgb
import optuna
from sklearn.metrics import confusion_matrix, classification_report, accuracy_score, precision_score, recall_score, f1_score, roc_curve
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import time
import openpyxl
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from scipy import optimize
from scipy.interpolate import interp1d
from feature_extractor import process_dataset_paths
import os
import threading

# признаки
FEATURE_GROUPS = {
    'RMS': 'rms',
    'ZCR': 'zcr',
    'Spectral Centroid': 'spec_cent',
    'Spectral Rolloff': 'spec_rolloff',
    'Spectral Flatness': 'spec_flatness',
    'MFCC': 'mfcc',
    'LFCC': 'lfcc',
    'Spectral Contrast': 'contrast'
}

MODELS = ['Random Forest', 'XGBoost', 'KNN']

class MLResearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ML Research Tool for Audio Deepfake Detection")
        self.root.geometry("900x700")

        # Create a canvas and scrollbar for the main window
        self.canvas = tk.Canvas(root)
        self.scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # скроллинг окна
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        # Variables
        self.training_path = tk.StringVar()
        self.validation_path = tk.StringVar()
        self.testing_path = tk.StringVar()
        # переменные для csv-файлов результатов извлечения признаков
        self.custom_train_csv = tk.StringVar()
        self.custom_valid_csv = tk.StringVar()
        self.custom_test_csv = tk.StringVar()
        self.selected_features = {}
        self.selected_model = tk.StringVar(value='Random Forest')
        self.hyper_ranges = {}
        self.num_trials = tk.StringVar(value='50')
        self.progress_var = tk.DoubleVar(value=0)
        self.status_text = tk.StringVar(value='Ready')
        self.threshold = 0.5
        self.best_params = None
        self.clf = None
        self.scaler = None
        self.extracted_csv_training = None
        self.extracted_csv_validation = None
        self.extracted_csv_testing = None
        # переменная директории сохранения результатов
        self.results_data = None
        # переменные времезатрат
        self.extraction_time = None
        self.optimization_time = None
        self.training_time = None
        self.feature_set_name = None
        # параметры извлечения признаков
        self.n_mfcc = tk.StringVar(value='20')
        self.n_lfcc = tk.StringVar(value='20')
        self.selected_stats = {}

        self.create_widgets()

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def create_widgets(self):
        # модуль извлечения признаков
        ttk.Label(self.scrollable_frame, text="Dataset Extraction", font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=2, pady=10)

        ttk.Label(self.scrollable_frame, text="Training Folder Path:").grid(row=1, column=0, sticky='w', padx=10, pady=5)
        train_frame = ttk.Frame(self.scrollable_frame)
        train_frame.grid(row=1, column=1, padx=10, pady=5)
        ttk.Entry(train_frame, textvariable=self.training_path, width=40).pack(side=tk.LEFT)
        ttk.Button(train_frame, text="Browse", command=lambda: self.browse_folder(self.training_path)).pack(side=tk.LEFT, padx=5)

        ttk.Label(self.scrollable_frame, text="Validation Folder Path:").grid(row=2, column=0, sticky='w', padx=10, pady=5)
        valid_frame = ttk.Frame(self.scrollable_frame)
        valid_frame.grid(row=2, column=1, padx=10, pady=5)
        ttk.Entry(valid_frame, textvariable=self.validation_path, width=40).pack(side=tk.LEFT)
        ttk.Button(valid_frame, text="Browse", command=lambda: self.browse_folder(self.validation_path)).pack(side=tk.LEFT, padx=5)

        ttk.Label(self.scrollable_frame, text="Testing Folder Path:").grid(row=3, column=0, sticky='w', padx=10, pady=5)
        test_frame = ttk.Frame(self.scrollable_frame)
        test_frame.grid(row=3, column=1, padx=10, pady=5)
        ttk.Entry(test_frame, textvariable=self.testing_path, width=40).pack(side=tk.LEFT)
        ttk.Button(test_frame, text="Browse", command=lambda: self.browse_folder(self.testing_path)).pack(side=tk.LEFT, padx=5)

        ttk.Label(self.scrollable_frame, text="Select Features to Extract:").grid(row=4, column=0, sticky='w', padx=10, pady=5)
        feature_frame = ttk.Frame(self.scrollable_frame)
        feature_frame.grid(row=5, column=0, columnspan=2, padx=10, pady=5)
        for i, group in enumerate(FEATURE_GROUPS.keys()):
            var = tk.BooleanVar(value=True)
            self.selected_features[group] = var
            ttk.Checkbutton(feature_frame, text=group, variable=var).grid(row=i//4, column=i%4, sticky='w')

        # количество коэффициентов MFCС/LFCC 
        ttk.Label(self.scrollable_frame, text="MFCC Coefficients:").grid(row=6, column=0, sticky='w', padx=10, pady=5)
        ttk.Entry(self.scrollable_frame, textvariable=self.n_mfcc, width=10).grid(row=6, column=1, sticky='w', padx=10, pady=5)
        
        ttk.Label(self.scrollable_frame, text="LFCC Coefficients:").grid(row=7, column=0, sticky='w', padx=10, pady=5)
        ttk.Entry(self.scrollable_frame, textvariable=self.n_lfcc, width=10).grid(row=7, column=1, sticky='w', padx=10, pady=5)

        # выбор статистических агрегатов
        ttk.Label(self.scrollable_frame, text="Select Statistics:").grid(row=8, column=0, sticky='w', padx=10, pady=5)
        stats_frame = ttk.Frame(self.scrollable_frame)
        stats_frame.grid(row=8, column=1, sticky='w', padx=10, pady=5)
        for stat in ['mean', 'std', 'skew', 'kurtosis']:
            var = tk.BooleanVar(value=True)
            self.selected_stats[stat] = var
            ttk.Checkbutton(stats_frame, text=stat, variable=var).pack(side=tk.LEFT, padx=5)

        ttk.Button(self.scrollable_frame, text="Extract Features and Save CSV", command=self.extract_features).grid(row=9, column=0, columnspan=2, pady=10)

        ttk.Button(self.scrollable_frame, text="Extract Features and Save CSV", command=self.extract_features).grid(row=9, column=0, columnspan=2, pady=10)

        # разделитель блоков
        ttk.Separator(self.scrollable_frame, orient='horizontal').grid(row=10, column=0, columnspan=2, sticky='ew', pady=10)

        ttk.Label(self.scrollable_frame, text="Load Custom Feature CSVs (Overrides Extraction)", font=('Arial', 12, 'bold')).grid(row=11, column=0, columnspan=2, pady=10)

        # поле ввода полного пути до CSV-файлов извлечения результатов набора Training
        ttk.Label(self.scrollable_frame, text="Custom Training CSV:").grid(row=12, column=0, sticky='w', padx=10, pady=5)
        ctrain_frame = ttk.Frame(self.scrollable_frame)
        ctrain_frame.grid(row=12, column=1, padx=10, pady=5)
        ttk.Entry(ctrain_frame, textvariable=self.custom_train_csv, width=40).pack(side=tk.LEFT)
        ttk.Button(ctrain_frame, text="Browse", command=lambda: self.browse_file(self.custom_train_csv)).pack(side=tk.LEFT, padx=5)

        # поле ввода полного пути до CSV-файлов извлечения результатов набора Validation
        ttk.Label(self.scrollable_frame, text="Custom Validation CSV:").grid(row=13, column=0, sticky='w', padx=10, pady=5)
        cvalid_frame = ttk.Frame(self.scrollable_frame)
        cvalid_frame.grid(row=13, column=1, padx=10, pady=5)
        ttk.Entry(cvalid_frame, textvariable=self.custom_valid_csv, width=40).pack(side=tk.LEFT)
        ttk.Button(cvalid_frame, text="Browse", command=lambda: self.browse_file(self.custom_valid_csv)).pack(side=tk.LEFT, padx=5)

        # поле ввода полного пути до CSV-файлов извлечения результатов набора Testing
        ttk.Label(self.scrollable_frame, text="Custom Testing CSV:").grid(row=14, column=0, sticky='w', padx=10, pady=5)
        ctest_frame = ttk.Frame(self.scrollable_frame)
        ctest_frame.grid(row=14, column=1, padx=10, pady=5)
        ttk.Entry(ctest_frame, textvariable=self.custom_test_csv, width=40).pack(side=tk.LEFT)
        ttk.Button(ctest_frame, text="Browse", command=lambda: self.browse_file(self.custom_test_csv)).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(self.scrollable_frame, orient='horizontal').grid(row=15, column=0, columnspan=2, sticky='ew', pady=10)

        # модели и гиперпараметры
        ttk.Label(self.scrollable_frame, text="Model Training", font=('Arial', 12, 'bold')).grid(row=16, column=0, columnspan=2, pady=10)

        ttk.Label(self.scrollable_frame, text="Select Model:").grid(row=17, column=0, sticky='w', padx=10, pady=5)
        model_combo = ttk.Combobox(self.scrollable_frame, textvariable=self.selected_model, values=MODELS)
        model_combo.grid(row=17, column=1, padx=10, pady=5)
        model_combo.bind("<<ComboboxSelected>>", self.update_hyper_entries)

        # диапазон перебора гиперпараметров
        self.hyper_frame = ttk.Frame(self.scrollable_frame)
        self.hyper_frame.grid(row=18, column=0, columnspan=2, padx=10, pady=5)
        self.update_hyper_entries()

        ttk.Label(self.scrollable_frame, text="Optuna Trials:").grid(row=19, column=0, sticky='w', padx=10, pady=5)
        ttk.Entry(self.scrollable_frame, textvariable=self.num_trials, width=12).grid(row=19, column=1, padx=10, pady=5)

        # кнопка запуска оптимизации гиперпараметров
        ttk.Button(self.scrollable_frame, text="Run Hyperparameter Optimization", command=self.run_optimization).grid(row=20, column=0, columnspan=2, pady=10)

        # линия прогресса
        self.progress_bar = ttk.Progressbar(self.scrollable_frame, maximum=100, variable=self.progress_var, length=600)
        self.progress_bar.grid(row=21, column=0, columnspan=2, padx=10, pady=5)
        ttk.Label(self.scrollable_frame, textvariable=self.status_text).grid(row=22, column=0, columnspan=2, sticky='w', padx=10, pady=2)

        # блок графика зависимости метрик от порогов вероятности
        self.plot_frame = ttk.Frame(self.scrollable_frame)
        self.plot_frame.grid(row=24, column=0, columnspan=2, padx=10, pady=10)

        # поле ввода порога вероятности
        ttk.Label(self.scrollable_frame, text="Selected Threshold:").grid(row=23, column=0, sticky='w', padx=10, pady=5)
        self.threshold_entry = ttk.Entry(self.scrollable_frame)
        self.threshold_entry.insert(0, "0.5")
        self.threshold_entry.grid(row=23, column=1, padx=10, pady=5)

        # кнопка запуска процесса обучения и тестирования модели
        ttk.Button(self.scrollable_frame, text="Run Final Evaluation", command=self.run_evaluation).grid(row=28, column=0, columnspan=2, pady=10)

        # блок вывода результатов
        results_frame = ttk.Frame(self.scrollable_frame)
        results_frame.grid(row=25, column=0, columnspan=2, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(results_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_text = tk.Text(results_frame, height=10, width=80, yscrollcommand=scrollbar.set)
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.results_text.yview)

        # кнопка сохранения результатов
        ttk.Button(self.scrollable_frame, text="Save Results to XLSX", command=self.save_results).grid(row=26, column=0, columnspan=2, pady=10)

    def browse_folder(self, target_var):
        folder = filedialog.askdirectory()
        if folder:
            target_var.set(folder)

    def extract_features(self): # функция запуска процесса извлечения признаков
        training_path = self.training_path.get()
        validation_path = self.validation_path.get()
        testing_path = self.testing_path.get()

        if not all([training_path, validation_path, testing_path]):
            messagebox.showerror("Error", "Provide paths for training, validation, and testing folders")
            return

        if not all(os.path.isdir(p) for p in [training_path, validation_path, testing_path]):
            messagebox.showerror("Error", "One or more split paths are invalid directories")
            return

        selected_groups = [FEATURE_GROUPS[group] for group, var in self.selected_features.items() if var.get()]
        if not selected_groups:
            messagebox.showerror("Error", "Select at least one feature group")
            return

        # прием выбранных статистических агрегатов
        stats_list = [stat for stat, var in self.selected_stats.items() if var.get()]
        if not stats_list:
            messagebox.showerror("Error", "Select at least one statistic")
            return

        # прием количества коэффициентов MFCC/LFCC
        try:
            n_mfcc = int(self.n_mfcc.get())
            n_lfcc = int(self.n_lfcc.get())
            if n_mfcc <= 0 or n_lfcc <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "MFCC and LFCC counts must be positive integers")
            return

        self.reset_progress()
        self.set_status("Extracting features...")
        
        # выключение кнопки на время извлечения признаков
        for child in self.scrollable_frame.winfo_children():
            if isinstance(child, ttk.Button) and child.cget('text') == 'Extract Features and Save CSV':
                child.config(state='disabled')
        
        # перевод процесса на параллельный поток во избежание зависания приложения
        def extraction_worker():
            try:
                start_time = time.time()
                output_paths = process_dataset_paths(
                    training_path,
                    validation_path,
                    testing_path,
                    feature_groups=selected_groups,
                    n_mfcc=n_mfcc,
                    n_lfcc=n_lfcc,
                    stats_list=stats_list,
                    progress_callback=self.on_extract_progress
                )
                self.extraction_time = time.time() - start_time
                
                # обновление линии прогресса в основном потоке
                self.root.after(0, self._extraction_complete, output_paths)
            except Exception as e:
                self.root.after(0, self._extraction_error, str(e))
        
        thread = threading.Thread(target=extraction_worker, daemon=True)
        thread.start()
    
    def _extraction_complete(self, output_paths):
        # активация кнопки запуска извлечения признаков
        for child in self.scrollable_frame.winfo_children():
            if isinstance(child, ttk.Button) and child.cget('text') == 'Extract Features and Save CSV':
                child.config(state='normal')
        
        self.extracted_csv_training = output_paths.get('training')
        self.extracted_csv_validation = output_paths.get('validation')
        self.extracted_csv_testing = output_paths.get('testing')
        self.feature_set_name = output_paths.get('feature_set_name')
        self.extraction_result_dir = output_paths.get('extraction_result_dir')

        if not all([self.extracted_csv_training, self.extracted_csv_validation, self.extracted_csv_testing]):
            messagebox.showerror("Error", 'Feature extraction did not produce all split CSV files')
            return

        self.set_progress(100, "Extraction complete")
        messagebox.showinfo("Success", f"Features extracted and saved to:\n- {self.extracted_csv_training}\n- {self.extracted_csv_validation}\n- {self.extracted_csv_testing}\n\nExtraction time: {self.extraction_time:.2f}s\nFeature set: {self.feature_set_name}")
    
    def _extraction_error(self, error_msg): # на случай возникновения ошибки в ходе извлечения признаков
        for child in self.scrollable_frame.winfo_children():
            if isinstance(child, ttk.Button) and child.cget('text') == 'Extract Features and Save CSV':
                child.config(state='normal')
        
        self.set_status("Extraction failed")
        messagebox.showerror("Error", error_msg)

    def update_hyper_entries(self, event=None):
        for widget in self.hyper_frame.winfo_children():
            widget.destroy()

        model = self.selected_model.get()
        if model == 'Random Forest':
            params = ['n_estimators', 'max_depth', 'min_samples_leaf']
        elif model == 'XGBoost':
            params = ['n_estimators', 'learning_rate', 'max_depth', 'gamma', 'lambda', 'alpha', 'subsample', 'colsample_bytree']
        elif model == 'KNN':
            params = ['n_neighbors']

        self.hyper_ranges = {}
        row = 0
        for param in params:
            ttk.Label(self.hyper_frame, text=f"{param} min:").grid(row=row, column=0, sticky='w')
            min_var = tk.StringVar(value='1' if param == 'n_neighbors' else '1')
            ttk.Entry(self.hyper_frame, textvariable=min_var, width=10).grid(row=row, column=1)
            ttk.Label(self.hyper_frame, text=f"{param} max:").grid(row=row, column=2, sticky='w')
            max_var = tk.StringVar(value='100' if param in ['n_estimators', 'max_depth'] else '20' if param == 'n_neighbors' else '1')
            ttk.Entry(self.hyper_frame, textvariable=max_var, width=10).grid(row=row, column=3)
            self.hyper_ranges[param] = (min_var, max_var)
            row += 1

    def load_data(self):
        # 1.  проверяем пользовательские CSV-файлы
        custom_train = self.custom_train_csv.get()
        custom_valid = self.custom_valid_csv.get()
        custom_test = self.custom_test_csv.get()

        if custom_train and custom_valid and custom_test:
            if all(os.path.exists(p) for p in [custom_train, custom_valid, custom_test]):
                self.set_status("Using custom user-provided feature CSVs")
                train_path = custom_train
                valid_path = custom_valid
                test_path = custom_test
            else:
                messagebox.showerror("Error", "One or more custom CSV paths are invalid or do not exist.") # если путь указан неверно
                return None, None, None, None, None, None
        else:
            # если результаты извлечения не указаны, будут задействованы последние извлеченные файлы результатов (во время текущей сессии)
            # в случае перезапуска приложения нужно обязательно указать директории .csv-файлов
            if hasattr(self, 'extraction_result_dir') and self.extraction_result_dir and os.path.exists(self.extraction_result_dir):
                extraction_csv_training = os.path.join(self.extraction_result_dir, 'training.csv')
                extraction_csv_validation = os.path.join(self.extraction_result_dir, 'validation.csv')
                extraction_csv_testing = os.path.join(self.extraction_result_dir, 'testing.csv')
                
                if all(os.path.exists(p) for p in [extraction_csv_training, extraction_csv_validation, extraction_csv_testing]):
                    self.extracted_csv_training = extraction_csv_training
                    self.extracted_csv_validation = extraction_csv_validation
                    self.extracted_csv_testing = extraction_csv_testing
                    self.set_status(f"Using extracted features from {self.extraction_result_dir}")
            
            paths = [self.extracted_csv_training, self.extracted_csv_validation, self.extracted_csv_testing]
            missing = [p for p in paths if not p or not os.path.exists(p)]
            if missing:
                self.try_infer_csv_paths()
                paths = [self.extracted_csv_training, self.extracted_csv_validation, self.extracted_csv_testing]
                missing = [p for p in paths if not p or not os.path.exists(p)]
                if missing:
                    messagebox.showerror("Error", "Provide custom CSVs OR extract features first to generate datasets.")
                    return None, None, None, None, None, None
            
            train_path = self.extracted_csv_training
            valid_path = self.extracted_csv_validation
            test_path = self.extracted_csv_testing

        # загрузка данных из определенных путей
        try:
            df_train = pd.read_csv(train_path)
            df_valid = pd.read_csv(valid_path)
            df_test = pd.read_csv(test_path)
            
            # bзвлечение признаков и меток
            X_train = df_train.drop(['filename', 'label', 'duration'], axis=1, errors='ignore')
            y_train = df_train['label']
            
            X_valid = df_valid.drop(['filename', 'label', 'duration'], axis=1, errors='ignore')
            y_valid = df_valid['label']
            
            X_test = df_test.drop(['filename', 'label', 'duration'], axis=1, errors='ignore')
            y_test = df_test['label']

            return X_train, X_valid, X_test, y_train, y_valid, y_test
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV files: {str(e)}")
            return None, None, None, None, None, None
    
    def browse_file(self, target_var):
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            target_var.set(file_path)

    def calculate_eer(self, y_true, y_score):
        fpr, tpr, thresholds = roc_curve(y_true, y_score, pos_label=1)
        eer = optimize.brentq(lambda x: 1. - x - interp1d(fpr, tpr)(x), 0., 1.)
        thresh = interp1d(fpr, thresholds)(eer)
        return eer, thresh

    def set_status(self, message):
        self.status_text.set(message)
        self.root.update_idletasks()

    def set_progress(self, value, message=None):
        self.progress_var.set(value)
        if message:
            self.status_text.set(message)
        self.root.update_idletasks()

    def reset_progress(self):
        self.progress_var.set(0)
        self.status_text.set('Ready')
        self.progress_bar.config(mode='determinate')
        self.root.update_idletasks()

    def on_extract_progress(self, percent, message=None):
        self.progress_var.set(percent)
        if message:
            self.status_text.set(message)
        self.root.update_idletasks()

    def infer_csv_path(self, split_name):
        candidate = None
        split_paths = [self.training_path.get(), self.validation_path.get(), self.testing_path.get()]
        valid_paths = [path for path in split_paths if path]
        if not valid_paths:
            return None
        try:
            output_dir = os.path.commonpath(valid_paths)
        except ValueError:
            output_dir = os.path.dirname(valid_paths[0])
        if os.path.isdir(output_dir):
            candidate = os.path.join(output_dir, f"{split_name}.csv")
            if os.path.exists(candidate):
                return candidate
        candidate = os.path.join(os.path.dirname(valid_paths[0]), f"{split_name}.csv")
        return candidate if os.path.exists(candidate) else None

    def try_infer_csv_paths(self):
        for split in ['training', 'validation', 'testing']:
            current = getattr(self, f"extracted_csv_{split}")
            if not current or not os.path.exists(current):
                inferred = self.infer_csv_path(split)
                if inferred:
                    setattr(self, f"extracted_csv_{split}", inferred)
                    self.set_status(f"Using inferred {split}.csv")

    def run_optimization(self):
        X_train, X_valid, X_test, y_train, y_valid, y_test = self.load_data()
        if X_train is None:
            return

        model = self.selected_model.get()

        if model == 'KNN':
            self.scaler = StandardScaler()
            X_train_scaled = self.scaler.fit_transform(X_train)
            X_valid_scaled = self.scaler.transform(X_valid)
            X_train_opt = X_train_scaled
            X_valid_opt = X_valid_scaled
        else:
            X_train_opt = X_train
            X_valid_opt = X_valid

        def objective(trial):
            if model == 'Random Forest':
                n_estimators = trial.suggest_int('n_estimators', int(self.hyper_ranges['n_estimators'][0].get()), int(self.hyper_ranges['n_estimators'][1].get()))
                max_depth = trial.suggest_int('max_depth', int(self.hyper_ranges['max_depth'][0].get()), int(self.hyper_ranges['max_depth'][1].get()))
                min_samples_leaf = trial.suggest_int('min_samples_leaf', int(self.hyper_ranges['min_samples_leaf'][0].get()), int(self.hyper_ranges['min_samples_leaf'][1].get()))
                clf = RandomForestClassifier(
                    n_estimators=n_estimators,
                    criterion='entropy',
                    max_depth=max_depth,
                    max_features='sqrt',
                    min_samples_leaf=min_samples_leaf,
                    random_state=42
                )
            elif model == 'XGBoost':
                params = {
                    'n_estimators': trial.suggest_int('n_estimators', int(self.hyper_ranges['n_estimators'][0].get()), int(self.hyper_ranges['n_estimators'][1].get())),
                    'learning_rate': trial.suggest_float('learning_rate', float(self.hyper_ranges['learning_rate'][0].get()), float(self.hyper_ranges['learning_rate'][1].get()), log=True),
                    'max_depth': trial.suggest_int('max_depth', int(self.hyper_ranges['max_depth'][0].get()), int(self.hyper_ranges['max_depth'][1].get())),
                    'gamma': trial.suggest_float('gamma', float(self.hyper_ranges['gamma'][0].get()), float(self.hyper_ranges['gamma'][1].get()), log=True),
                    'lambda': trial.suggest_float('lambda', float(self.hyper_ranges['lambda'][0].get()), float(self.hyper_ranges['lambda'][1].get()), log=True),
                    'alpha': trial.suggest_float('alpha', float(self.hyper_ranges['alpha'][0].get()), float(self.hyper_ranges['alpha'][1].get()), log=True),
                    'subsample': trial.suggest_float('subsample', float(self.hyper_ranges['subsample'][0].get()), float(self.hyper_ranges['subsample'][1].get())),
                    'colsample_bytree': trial.suggest_float('colsample_bytree', float(self.hyper_ranges['colsample_bytree'][0].get()), float(self.hyper_ranges['colsample_bytree'][1].get())),
                    'random_state': 42
                }
                clf = xgb.XGBClassifier(**params)
            elif model == 'KNN':
                n_neighbors = trial.suggest_int('n_neighbors', int(self.hyper_ranges['n_neighbors'][0].get()), int(self.hyper_ranges['n_neighbors'][1].get()))
                weights = trial.suggest_categorical('weights', ['uniform', 'distance'])
                clf = KNeighborsClassifier(n_neighbors=n_neighbors, weights=weights)

            clf.fit(X_train_opt, y_train)
            if hasattr(clf, 'predict_proba'):
                y_score = clf.predict_proba(X_valid_opt)[:, 1]
            else:
                y_score = clf.predict(X_valid_opt)  # For KNN without proba

            # Use EER for all models (better and consistent)
            eer, best_threshold = self.calculate_eer(y_valid, y_score)
            trial.set_user_attr("eer_threshold", float(best_threshold))
            trial.set_user_attr("accuracy", accuracy_score(y_valid, y_score > 0.5))
            return eer

        try:
            trials = int(self.num_trials.get())
            if trials <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Enter a valid positive integer for Optuna trials")
            return

        study = optuna.create_study(direction="minimize")

        def optuna_progress_callback(study_, trial):
            index = trial.number + 1
            self.set_progress(index / trials * 100, f"Optimization trial {index} / {trials}")

        opt_start_time = time.time()
        study.optimize(objective, n_trials=trials, callbacks=[optuna_progress_callback])
        self.optimization_time = time.time() - opt_start_time

        self.best_params = study.best_params
        self.best_threshold = study.best_trial.user_attrs.get("eer_threshold", 0.5)
        self.set_progress(100, "Optimization complete")

        messagebox.showinfo("Optimization Complete", f"Best params: {self.best_params}\n\nOptimization time: {self.optimization_time:.2f}s")

        # запуск обучения и оценки на тестовых данных
        self.run_evaluation()

    def run_validation_evaluation(self):
        X_train, X_valid, X_test, y_train, y_valid, y_test = self.load_data()
        if X_train is None:
            return

        model = self.selected_model.get()

        # обучение с валидационным сетом
        if model == 'KNN':
            if self.scaler is None:
                self.scaler = StandardScaler()
            X_valid_scaled = self.scaler.fit_transform(X_valid)
            X_eval = X_valid_scaled
        else:
            X_eval = X_valid

        # обучение на тренировочном сете с лучшими гиперпараметрами
        if self.best_params is None:
            messagebox.showerror("Error", "Run optimization first")
            return

        self.progress_bar.config(mode='indeterminate')
        self.progress_bar.start(10)
        self.set_status("Training model for validation evaluation...")
        start_time = time.time()
        if model == 'Random Forest':
            clf = RandomForestClassifier(**self.best_params, random_state=42)
        elif model == 'XGBoost':
            clf = xgb.XGBClassifier(**self.best_params)
        elif model == 'KNN':
            clf = KNeighborsClassifier(**self.best_params)
        clf.fit(X_train if model != 'KNN' else self.scaler.transform(X_train), y_train)
        training_time = time.time() - start_time
        self.progress_bar.stop()
        self.progress_bar.config(mode='determinate')
        self.set_status("Model trained for validation")

        # предсказание на валидационных данных
        start_time = time.time()
        if hasattr(clf, 'predict_proba'):
            y_pred_proba = clf.predict_proba(X_eval)[:, 1]
            y_pred = (y_pred_proba > 0.5).astype(int)
        else:
            y_pred = clf.predict(X_eval)
            y_pred_proba = y_pred
        prediction_time = time.time() - start_time

        # вывод графика зависимости метрик от порога вероятности (в случае, если не указан тестовый набор)
        if hasattr(clf, 'predict_proba'):
            self.plot_threshold_metrics(y_valid, y_pred_proba)
        else:
            self.threshold = 0.5

        # вывод результатов оценки на валидационном сете с порогом по умолчанию 0.5
        acc = accuracy_score(y_valid, y_pred)
        prec = precision_score(y_valid, y_pred)
        rec = recall_score(y_valid, y_pred)
        f1 = f1_score(y_valid, y_pred)

        cm = confusion_matrix(y_valid, y_pred)

        results = f"Validation Evaluation (Default Threshold 0.5):\n"
        results += f"Training Time: {training_time:.2f}s\nPrediction Time: {prediction_time:.2f}s\n\n"
        results += f"Accuracy: {acc:.3f}\nPrecision: {prec:.3f}\nRecall: {rec:.3f}\nF1-Score: {f1:.3f}\n\n"
        results += f"Confusion Matrix:\n{cm}\n\n"
        results += "Adjust threshold above and run Final Evaluation on test set."

        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, results)

    def run_evaluation(self):
        X_train, X_valid, X_test, y_train, y_valid, y_test = self.load_data()
        if X_train is None:
            return

        model = self.selected_model.get()

        # слияние валидационного и тренировочного наборов
        X_train_valid = pd.concat([X_train, X_valid], ignore_index=True)
        y_train_valid = pd.concat([y_train, y_valid], ignore_index=True)

        if model == 'KNN':
            if self.scaler is None:
                self.scaler = StandardScaler()
            X_train_valid_scaled = self.scaler.fit_transform(X_train_valid)
            X_test_scaled = self.scaler.transform(X_test)
            X_eval = X_train_valid_scaled
            X_test_eval = X_test_scaled
        else:
            X_eval = X_train_valid
            X_test_eval = X_test

        # обучение с лушими гиперпараметрами
        if self.best_params is None:
            messagebox.showerror("Error", "Run optimization first")
            return

        self.progress_bar.config(mode='indeterminate')
        self.progress_bar.start(10)
        self.set_status("Training final model...")
        start_time = time.time()
        if model == 'Random Forest':
            self.clf = RandomForestClassifier(**self.best_params, random_state=42)
        elif model == 'XGBoost':
            self.clf = xgb.XGBClassifier(**self.best_params)
        elif model == 'KNN':
            self.clf = KNeighborsClassifier(**self.best_params)
        self.clf.fit(X_eval, y_train_valid)
        self.training_time = time.time() - start_time
        training_time = self.training_time
        self.progress_bar.stop()
        self.progress_bar.config(mode='determinate')
        self.set_status("Model trained")

        # предсказание результатов с порогом вероятности 0.5
        start_time = time.time()
        if hasattr(self.clf, 'predict_proba'):
            y_pred_proba = self.clf.predict_proba(X_test_eval)[:, 1]
            y_pred = (y_pred_proba > 0.5).astype(int)
        else:
            y_pred = self.clf.predict(X_test_eval)
            y_pred_proba = y_pred
        prediction_time = time.time() - start_time

        # вывод графика зависимости тестовых метрик от порога вероятности
        if hasattr(self.clf, 'predict_proba'):
            self.plot_threshold_metrics(y_test, y_pred_proba)
        else:
            self.threshold = 0.5

        # выбор порога вероятности
        try:
            self.threshold = float(self.threshold_entry.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid threshold value")
            return

        # рассчет метрик с учетом указанного порога верояности
        if hasattr(self.clf, 'predict_proba'):
            y_pred = (y_pred_proba > self.threshold).astype(int)

        # метрики
        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred)
        rec = recall_score(y_test, y_pred)
        f1 = f1_score(y_test, y_pred)

        cm = confusion_matrix(y_test, y_pred)

        # рассчет важности признаков
        if model in ['Random Forest', 'XGBoost']:
            if hasattr(self.clf, 'feature_importances_'):
                importances = self.clf.feature_importances_
                feature_names = X_train.columns
            else:
                importances = None
        else:
            importances = None

        # вывод результатов
        results = f"=== Timing Information ===\n"
        # results += f"Feature Extraction Time: {self.extraction_time:.2f}s\n"
        results += f"Hyperparameter Optimization Time: {self.optimization_time:.2f}s\n"
        results += f"Model Training Time: {training_time:.2f}s\n"
        results += f"Prediction Time: {prediction_time:.2f}s\n\n"
        results += f"=== Metrics ===\n"
        results += f"Accuracy: {acc:.3f}\nPrecision: {prec:.3f}\nRecall: {rec:.3f}\nF1-Score: {f1:.3f}\n\n"
        results += f"Confusion Matrix:\n"
        results += f"{'':<10} {'Actual Real':<12} {'Actual Fake':<12}\n"
        results += f"{'Pred Real':<10} {cm[0,0]:<12} {cm[0,1]:<12}\n"
        results += f"{'Pred Fake':<10} {cm[1,0]:<12} {cm[1,1]:<12}\n\n"

        if importances is not None:
            results += "Feature Importances:\n"
            for name, imp in zip(feature_names, importances):
                results += f"{name}: {imp:.3f}\n"

        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, results)

        # результаты для сохранения
        self.results_data = {
            'acc': acc,
            'prec': prec,
            'rec': rec,
            'f1': f1,
            'training_time': training_time,
            'prediction_time': prediction_time,
            'extraction_time': self.extraction_time,
            'optimization_time': self.optimization_time,
            'feature_set_name': self.feature_set_name,
            'cm': cm,
            'importances': importances,
            'feature_names': feature_names if importances is not None else None,
            'best_params': self.best_params,
            'model': model
        }

    def plot_threshold_metrics(self, y_true, y_proba):
        thresholds = np.arange(0.0, 1.01, 0.01)
        precisions = []
        recalls = []
        f1s = []
        accuracies = []

        for thresh in thresholds:
            y_pred = (y_proba > thresh).astype(int)
            precisions.append(precision_score(y_true, y_pred, zero_division=0))
            recalls.append(recall_score(y_true, y_pred, zero_division=0))
            f1s.append(f1_score(y_true, y_pred, zero_division=0))
            accuracies.append(accuracy_score(y_true, y_pred))

        fig, ax = plt.subplots(figsize=(8, 6))
        ax.plot(thresholds, precisions, label='Precision')
        ax.plot(thresholds, recalls, label='Recall')
        ax.plot(thresholds, f1s, label='F1-Score')
        ax.plot(thresholds, accuracies, label='Accuracy')
        ax.set_xlabel('Threshold')
        ax.set_ylabel('Score')
        ax.set_title('Metrics vs Threshold')
        ax.legend()
        ax.grid(True)

        # очистка графика
        for widget in self.plot_frame.winfo_children():
            widget.destroy()

        canvas = FigureCanvasTkAgg(fig, master=self.plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def save_to_xlsx(self, data):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        # Набор данных
        ws['A1'] = 'Feature Set Name'
        ws['B1'] = data.get('feature_set_name', 'N/A')
        ws['A2'] = 'Model'
        ws['B2'] = data.get('model', 'N/A')
        ws['A3'] = 'Best Parameters'
        ws['B3'] = str(data.get('best_params', {}))

        # Времязатраты
        ws['A5'] = 'Timing Information'
        ws['B5'] = 'Time (s)'
        # ws['A6'] = 'Feature Extraction Time'
        ws['B6'] = data.get('extraction_time', 0)
        ws['A7'] = 'Hyperparameter Optimization Time'
        ws['B7'] = data.get('optimization_time', 0)
        ws['A8'] = 'Model Training Time'
        ws['B8'] = data.get('training_time', 0)
        ws['A9'] = 'Prediction Time'
        ws['B9'] = data.get('prediction_time', 0)

        # Метрики
        ws['A11'] = 'Metric'
        ws['B11'] = 'Value'
        ws['A12'] = 'Accuracy'
        ws['B12'] = data.get('acc', 0)
        ws['A13'] = 'Precision'
        ws['B13'] = data.get('prec', 0)
        ws['A14'] = 'Recall'
        ws['B14'] = data.get('rec', 0)
        ws['A15'] = 'F1-Score'
        ws['B15'] = data.get('f1', 0)

        # Confusion Matrix
        ws['D11'] = 'Confusion Matrix'
        cm = data.get('cm')
        if cm is not None:
            for i in range(cm.shape[0]):
                for j in range(cm.shape[1]):
                    ws.cell(row=i+12, column=j+4, value=cm[i,j])

        # важность признаков
        importances = data.get('importances')
        feature_names = data.get('feature_names')
        if importances is not None and feature_names is not None:
            ws['A18'] = 'Feature'
            ws['B18'] = 'Importance'
            for i, (name, imp) in enumerate(zip(feature_names, importances)):
                ws.cell(row=i+19, column=1, value=name)
                ws.cell(row=i+19, column=2, value=imp)

        wb.save(file_path)
        messagebox.showinfo("Saved", f"Results saved to {file_path}")

    def save_results(self):
        if self.results_data is None:
            messagebox.showerror("Error", "Run final evaluation first to generate results")
            return
        self.save_to_xlsx(self.results_data)

if __name__ == "__main__":
    root = tk.Tk()
    app = MLResearchApp(root)
    root.mainloop()